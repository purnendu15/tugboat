# Copyright 2018 AT&T Intellectual Property.  All other rights reserved.
#
# Licensed under the Apache License, Version 2.0 (the "License");
# you may not use this file except in compliance with the License.
# You may obtain a copy of the License at
#
# http://www.apache.org/licenses/LICENSE-2.0
#
# Unless required by applicable law or agreed to in writing, software
# distributed under the License is distributed on an "AS IS" BASIS,
# WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
# See the License for the specific language governing permissions and
# limitations under the License.

import json
import jsonschema
import pkg_resources
import re
import sys
import yaml
from openpyxl import load_workbook

from ..check_exceptions import (
    NoSpecMatched, )
import logging
import pprint


class ExcelParser():
    """ Parse data from excel into a dict """

    def __init__(self, file_name, excel_specs):
        self.logger = logging.getLogger(__name__)
        self.file_name = file_name
        with open(excel_specs, 'r') as f:
            spec_raw_data = f.read()
        self.excel_specs = yaml.safe_load(spec_raw_data)
        self.wb = load_workbook(file_name, data_only=True)
        self.spec = None

    @staticmethod
    def sanitize(string):
        """ Remove extra spaces and convert string to lower case """
        return string.replace(' ', '').lower()

    def compare(self, string1, string2):
        """ Compare the strings """
        return bool(re.search(self.sanitize(string1), self.sanitize(string2)))

    def validate_sheet(self, spec, sheet):
        """ Check if the sheet is correct or not """
        ws = self.wb[sheet]
        header_row = self.excel_specs['specs'][spec]['header_row']
        ipmi_header = self.excel_specs['specs'][spec]['ipmi_address_header']
        ipmi_column = self.excel_specs['specs'][spec]['ipmi_address_col']
        header_value = ws.cell(row=header_row, column=ipmi_column).value
        return bool(self.compare(ipmi_header, header_value))

    def find_correct_spec(self):
        """ Find the correct spec """
        for spec in self.excel_specs['specs']:
            sheet_name = self.excel_specs['specs'][spec]['ipmi_sheet_name']
            for sheet in self.wb.sheetnames:
                if self.compare(sheet_name, sheet):
                    self.excel_specs['specs'][spec]['ipmi_sheet_name'] = sheet
                    if self.validate_sheet(spec, sheet):
                        return spec
        raise NoSpecMatched(self.excel_specs)

    def get_ipmi_data(self):
        """ Read IPMI data from the sheet """
        ipmi_data = {}
        hosts = []
        self.spec = self.find_correct_spec()
        sheet_name = self.excel_specs['specs'][self.spec]['ipmi_sheet_name']
        ws = self.wb[sheet_name]
        row = self.excel_specs['specs'][self.spec]['start_row']
        end_row = self.excel_specs['specs'][self.spec]['end_row']
        hostname_col = self.excel_specs['specs'][self.spec]['hostname_col']
        ipmi_address_col = self.excel_specs['specs'][self.spec][
            'ipmi_address_col']
        host_profile_col = self.excel_specs['specs'][self.spec][
            'host_profile_col']
        ipmi_gateway_col = self.excel_specs['specs'][self.spec][
            'ipmi_gateway_col']
        previous_server_gateway = None
        while row <= end_row:
            hostname = self.sanitize(
                ws.cell(row=row, column=hostname_col).value)
            hosts.append(hostname)
            ipmi_address = ws.cell(row=row, column=ipmi_address_col).value
            if '/' in ipmi_address:
                ipmi_address = ipmi_address.split('/')[0]
            ipmi_gateway = ws.cell(row=row, column=ipmi_gateway_col).value
            if ipmi_gateway:
                previous_server_gateway = ipmi_gateway
            else:
                ipmi_gateway = previous_server_gateway
            tmp_host_profile = ws.cell(row=row, column=host_profile_col).value
            try:
                if tmp_host_profile is None:
                    raise RuntimeError("No value read from {} ".format(
                        self.file_name) + "sheet:{} row:{}, col:{}".format(
                                           self.spec, row, host_profile_col))
            except RuntimeError as rerror:
               self.logger.critical(rerror)
               sys.exit("Tugboat existed!!")
            host_profile = tmp_host_profile.split('-')[1]
            
            ipmi_data[hostname] = {
                'ipmi_address': ipmi_address,
                'ipmi_gateway': ipmi_gateway,
                'host_profile': host_profile,
            }
            row += 1
        self.logger.debug("ipmi data extracted from excel:\n%s",
                          [pprint.pformat(ipmi_data),
                           pprint.pformat(hosts)])
        return [ipmi_data, hosts]

    def get_private_vlan_data(self, ws):
        """ Get private vlan data from private IP sheet """
        vlan_data = {}
        row = self.excel_specs['specs'][self.spec]['vlan_start_row']
        end_row = self.excel_specs['specs'][self.spec]['vlan_end_row']
        type_col = self.excel_specs['specs'][self.spec]['net_type_col']
        vlan_col = self.excel_specs['specs'][self.spec]['vlan_col']
        while row <= end_row:
            cell_value = ws.cell(row=row, column=type_col).value
            if cell_value:
                vlan = ws.cell(row=row, column=vlan_col).value
                if vlan:
                    vlan = vlan.lower()
                vlan_data[vlan] = cell_value
            row += 1
        self.logger.debug("vlan data extracted from excel:\n%s",
                          pprint.pformat(vlan_data))
        return vlan_data

    def get_private_network_data(self):
        """ Read network data from the private ip sheet """
        sheet_name = self.excel_specs['specs'][self.spec]['private_ip_sheet']
        ws = self.wb[sheet_name]
        vlan_data = self.get_private_vlan_data(ws)
        network_data = {}
        row = self.excel_specs['specs'][self.spec]['net_start_row']
        end_row = self.excel_specs['specs'][self.spec]['net_end_row']
        col = self.excel_specs['specs'][self.spec]['net_col']
        vlan_col = self.excel_specs['specs'][self.spec]['net_vlan_col']
        old_vlan = ''
        while row <= end_row:
            vlan = ws.cell(row=row, column=vlan_col).value
            if vlan:
                vlan = vlan.lower()
            network = ws.cell(row=row, column=col).value
            if vlan and network:
                net_type = vlan_data[vlan]
                if 'vlan' not in network_data:
                    network_data[net_type] = {
                        'vlan': vlan,
                        'subnet': [],
                    }
            elif not vlan and network:
                # If vlan is not present then assign old vlan to vlan as vlan
                # value is spread over several rows
                vlan = old_vlan
            else:
                row += 1
                continue
            network_data[vlan_data[vlan]]['subnet'].append(network)
            old_vlan = vlan
            row += 1
        for network in network_data:
            if len(network_data[network]['subnet']) > 1:
                network_data[network]['is_common'] = False
            else:
                network_data[network]['is_common'] = True
        self.logger.debug(
            "private network data extracted from\
                          excel:\n%s", pprint.pformat(network_data))
        return network_data

    def get_public_network_data(self):
        """ Read public network data from public ip data """
        network_data = {}
        sheet_name = self.excel_specs['specs'][self.spec]['public_ip_sheet']
        ws = self.wb[sheet_name]
        oam_row = self.excel_specs['specs'][self.spec]['oam_ip_row']
        oam_col = self.excel_specs['specs'][self.spec]['oam_ip_col']
        oam_vlan_col = self.excel_specs['specs'][self.spec]['oam_vlan_col']
        ingress_row = self.excel_specs['specs'][self.spec]['ingress_ip_row']
        oob_row = self.excel_specs['specs'][self.spec]['oob_net_row']
        col = self.excel_specs['specs'][self.spec]['oob_net_start_col']
        end_col = self.excel_specs['specs'][self.spec]['oob_net_end_col']
        network_data = {
            'oam': {
                'ip': ws.cell(row=oam_row, column=oam_col).value,
                'vlan': ws.cell(row=oam_row, column=oam_vlan_col).value,
            },
            'ingress': ws.cell(row=ingress_row, column=oam_col).value,
        }
        network_data['oob'] = {
            'subnets': [],
        }
        while col <= end_col:
            cell_value = ws.cell(row=oob_row, column=col).value
            if cell_value:
                network_data['oob']['subnets'].append(
                    self.sanitize(cell_value))
            col += 1
        self.logger.debug(
            "public network data extracted from\
                          excel:\n%s", pprint.pformat(network_data))
        return network_data

    def get_dns_ntp_ldap_data(self):
        """ Read dns, ntp and ldap data from build notes sheet """
        dns_ntp_ldap_data = {}
        sheet_name = self.excel_specs['specs'][self.spec]['dns_ntp_ldap_sheet']
        ws = self.wb[sheet_name]
        dns_row = self.excel_specs['specs'][self.spec]['dns_row']
        dns_col = self.excel_specs['specs'][self.spec]['dns_col']
        ntp_row = self.excel_specs['specs'][self.spec]['ntp_row']
        ntp_col = self.excel_specs['specs'][self.spec]['ntp_col']
        domain_row = self.excel_specs['specs'][self.spec]['domain_row']
        domain_col = self.excel_specs['specs'][self.spec]['domain_col']
        ldap_subdomain_row = self.excel_specs['specs'][self.spec][
            'ldap_subdomain_row']
        ldap_col = self.excel_specs['specs'][self.spec]['ldap_col']
        ldap_group_row = self.excel_specs['specs'][self.spec]['ldap_group_row']
        ldap_url_row = self.excel_specs['specs'][self.spec]['ldap_url_row']
        dns_servers = ws.cell(row=dns_row, column=dns_col).value
        ntp_servers = ws.cell(row=ntp_row, column=ntp_col).value
        try:
            if dns_servers is None:
                raise RuntimeError("No value read for dns_server from File:" +
                    "{} Sheet:'{}' Row:{} Col:{}".format(
                        self.file_name, sheet_name,
                                                  dns_row, dns_col))
                raise RuntimeError("No value read for ntp_server from File:" +
                    "{} Sheet:'{}' Row:{} Col:{}".format(
                        self.file_name, sheet_name,
                                                  ntp_row, ntp_col))
        except RuntimeError as rerror:
            self.logger.critical(rerror)
            sys.exit("Tugboat existed!!")

        dns_servers = dns_servers.replace('\n', ' ')
        ntp_servers = ntp_servers.replace('\n', ' ')
        if ',' in dns_servers:
            dns_servers = dns_servers.split(',')
        else:
            dns_servers = dns_servers.split()
        if ',' in ntp_servers:
            ntp_servers = ntp_servers.split(',')
        else:
            ntp_servers = ntp_servers.split()
        dns_ntp_ldap_data = {
            'dns': dns_servers,
            'ntp': ntp_servers,
            'domain': ws.cell(row=domain_row, column=domain_col).value,
            'ldap': {
                'subdomain':
                ws.cell(row=ldap_subdomain_row, column=ldap_col).value,
                'common_name':
                ws.cell(row=ldap_group_row, column=ldap_col).value,
                'url':
                ws.cell(row=ldap_url_row, column=ldap_col).value,
            }
        }
        self.logger.debug(
            "dns,ntp,ldap data extracted from\
                          excel:\n%s", pprint.pformat(dns_ntp_ldap_data))
        return dns_ntp_ldap_data

    def get_location_data(self):
        """ Read location data from the site and zone sheet """
        sheet_name = self.excel_specs['specs'][self.spec]['location_sheet']
        ws = self.wb[sheet_name]
        corridor_row = self.excel_specs['specs'][self.spec]['corridor_row']
        column = self.excel_specs['specs'][self.spec]['column']
        name_row = self.excel_specs['specs'][self.spec]['name_row']
        state_row = self.excel_specs['specs'][self.spec]['state_row']
        country_row = self.excel_specs['specs'][self.spec]['country_row']
        clli_row = self.excel_specs['specs'][self.spec]['clli_row']
        return {
            'corridor': ws.cell(row=corridor_row, column=column).value,
            'name': ws.cell(row=name_row, column=column).value,
            'state': ws.cell(row=state_row, column=column).value,
            'country': ws.cell(row=country_row, column=column).value,
            'physical_location_id': ws.cell(row=clli_row, column=column).value,
        }

    def validate_data(self, data):
        self.logger.info('Validating data read from sheet')
        schema_dir = pkg_resources.resource_filename('tugboat', 'schemas/')
        schema_file = schema_dir + "data_schema.json"
        json_data = json.loads(json.dumps(data))
        with open(schema_file, 'r') as f:
            json_schema = json.load(f)
        try:
            with open('data2.json', 'w') as outfile:
                json.dump(data,outfile,sort_keys=True, indent=4)
            jsonschema.validate(json_data, json_schema)
        except jsonschema.exceptions.ValidationError as e:
            self.logger.error(
                "Validation Failed with following error:i" +
                "\n{}\n Please check excel spec settings(row,col)".format(
                    e.message
                )
            )
            sys.exit(1)
        self.logger.info("Data validation\
                         OK!")

    def validate_sheet_names_with_spec(self):
        spec = list(self.excel_specs['specs'].keys())[0]
        spec_item = self.excel_specs['specs'][spec]
        ipmi_header_sheet_name = spec_item['ipmi_sheet_name']
        private_ip_sheet_name = spec_item['private_ip_sheet']
        public_ip_sheet_name = spec_item['public_ip_sheet']
        dns_ntp_ldap_sheet_name = spec_item['dns_ntp_ldap_sheet']
        location_sheet_name = spec_item['location_sheet']

        try:
            if ipmi_header_sheet_name not in self.wb.sheetnames:
                raise RuntimeError(
                    "SheetName '{}' not found in '{}'".format(
                        ipmi_header_sheet_name, self.file_name))
            if private_ip_sheet_name not in self.wb.sheetnames:
                raise RuntimeError(
                    "SheetName '{}' not found in '{}'".format(
                        private_ip_sheet_name, self.file_name))
            if public_ip_sheet_name not in self.wb.sheetnames:
                raise RuntimeError(
                    "SheetName '{}' not found in '{}'".format(
                        public_ip_sheet_name, self.file_name))
            if dns_ntp_ldap_sheet_name not in self.wb.sheetnames:
                raise RuntimeError(
                    "SheetName '{}' not found in '{}'".format(
                        dns_ntp_ldap_sheet_name, self.file_name))
            if location_sheet_name not in self.wb.sheetnames:
                raise RuntimeError(
                    "SheetName '{}' not found in '{}'".format(
                        location_sheet_name, self.file_name))
        except RuntimeError as rerror:
            self.logger.critical(rerror)
            sys.exit("Tugboat exited!!")
        self.logger.info(
            "Sheet name in excel spec validated with'{}'".format(
                self.file_name))

    def get_data(self):
        """ Create a dict with combined data """
        self.validate_sheet_names_with_spec()
        ipmi_data = self.get_ipmi_data()
        network_data = self.get_private_network_data()
        public_network_data = self.get_public_network_data()
        dns_ntp_ldap_data = self.get_dns_ntp_ldap_data()
        location_data = self.get_location_data()
        self.logger.debug(
            "Location data extracted from\
                          excel:\n%s", pprint.pformat(location_data))
        data = {
            'ipmi_data': ipmi_data,
            'network_data': {
                'private': network_data,
                'public': public_network_data,
                'dns_ntp_ldap': dns_ntp_ldap_data,
            },
            'location_data': location_data,
        }
        self.validate_data(data)
        return data
